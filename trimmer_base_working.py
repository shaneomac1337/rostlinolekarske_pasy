from openpyxl import load_workbook

# Load the Excel file
excel_file = '23092023.xlsx'  # replace with your Excel file name

# Load the dictionary.txt file
with open('dictionary.txt', 'r') as f:
    dictionary_values = f.read().splitlines()

# Load the workbook
book = load_workbook(excel_file)

# Iterate over all sheets in the workbook
for sheet_name in book.sheetnames:
    sheet = book[sheet_name]

    # Iterate over the cells in column 'C' starting from row 13
    for row in range(13, sheet.max_row + 1):
        cell = sheet.cell(row=row, column=3)
        if cell.value:  # check if the cell has a value
            matched = False
            for value in dictionary_values:
                if value in cell.value:
                    print(f'Matched: {cell.value} -> {value}')
                    cell.value = value
                    matched = True
                    break
            if not matched:
                print(f'Unmatched: {cell.value}')

# Save the workbook
book.save(excel_file)

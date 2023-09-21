import os
import openpyxl
import pandas as pd
from copy import copy

# Path to the directory containing the script and the mail_tool folder
dir_path = os.path.dirname(os.path.realpath(__file__))

# Read the Excel files
recipients_df = pd.read_excel(os.path.join(dir_path, 'mail_tool', 'recipients.xlsx'))
build_wb = openpyxl.load_workbook(os.path.join(dir_path, 'mail_tool', 'build.xlsx'))

# Get the unique values in the second column of the recipients file, excluding the '.pdf' part
unique_values = recipients_df.iloc[:, 1].dropna().unique()

# Check if there are any unique values
if len(unique_values) == 0:
    print("No unique values found in the second column of the recipients file.")
else:
    # Ask for the name of the new Excel file
    new_file_name = input("Enter the name for the new Excel file: ")

    # Add the '.xlsx' suffix if it's not already there
    if not new_file_name.endswith('.xlsx'):
        new_file_name += '.xlsx'

    # Create a new Excel workbook
    new_wb = openpyxl.Workbook()
    new_wb.remove(new_wb.active)  # remove the default sheet created

    # Loop over the unique values
    for value in unique_values:
        sheet_name = str(value).replace('.pdf', '')
        # Copy the data from the build file to a new sheet in the new Excel file
        source = build_wb.active
        target = new_wb.create_sheet(title=sheet_name)

        for row in source:
            for cell in row:
                new_cell = target.cell(row=cell.row, column=cell.column, value=cell.value)
                if cell.has_style:
                    new_cell.font = copy(cell.font)
                    new_cell.border = copy(cell.border)
                    new_cell.fill = copy(cell.fill)
                    new_cell.number_format = copy(cell.number_format)
                    new_cell.protection = copy(cell.protection)
                    new_cell.alignment = copy(cell.alignment)

        # Copy column widths
        for column in source.columns:
            max_length = 0
            column = [cell for cell in column]
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = (max_length + 2)
            target.column_dimensions[column[0].column_letter].width = adjusted_width

        # Copy row heights
        for row in source.rows:
            for cell in row:
                target.row_dimensions[cell.row].height = source.row_dimensions[cell.row].height

    # Save the new Excel file
    new_wb.save(os.path.join(dir_path, new_file_name))

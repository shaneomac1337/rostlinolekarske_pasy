import os
import sys
import tkinter as tk
from tkinter import ttk
from tkinter import messagebox
from tkinter import filedialog
import openpyxl
from fuzzywuzzy import fuzz
import win32com.client
import requests
import semver
import webbrowser
import win32com.client as win32
from openpyxl.styles import PatternFill
import pandas as pd
from openpyxl import load_workbook
import time
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font, Border, Side, Alignment
from PyPDF2 import PdfReader
import re
from copy import copy
from tkinter import simpledialog
from openpyxl.utils import get_column_letter
import shutil
from tkinter import ttk, Toplevel, Text, Button, END, messagebox
from openpyxl import load_workbook
import gc
from openpyxl import Workbook
from openpyxl.styles import NamedStyle
import webbrowser
import threading
import queue
import pythoncom
from PyPDF2 import PdfMerger, PdfReader, PdfWriter


current_version = "v1.2.9"
url = 'https://api.github.com/repos/{owner}/{repo}/releases/latest'

try:
    response = requests.get(url.format(owner='shaneomac1337', repo='rostlinolekarske_pasy'))

    if response.status_code == requests.codes.ok:
        latest_release = response.json()
        latest_version = latest_release['tag_name'][1:]
        if semver.compare(current_version[1:], latest_version) < 0:
            # Display a message box to the user
            app = tk.Tk()
            app.withdraw()
            result = messagebox.askyesno('Aktualizace dostupná', 'Nová verze toolu na pasy je k dispozici, přeje si Olinka stáhnout novou verzi z webu?')

            if result:
                # Otevřít GitHub k nalezení aktuální verze
                url = latest_release['html_url']
                webbrowser.open_new(url)

                # Update je k dispozici, detekce platformy
                assets = latest_release['assets']
                for asset in assets:
                    if 'Windows' in asset['name'] and 'x86_64' in asset['name']:
                        download_url = asset['browser_download_url']
                        r = requests.get(download_url)
            else:
                # Nedělat nic, pokud zvoleno "Ne"
                pass
        else:
            # Update není k dispozici
            pass
    else:
        # Neobdržel jsem info o updatu
        pass
except requests.exceptions.RequestException:
    # Display an error message box
    messagebox.showerror("Chyba", "Olince nejde internet, copa s tim vyvádí?? No každopádně tool se pustí i bez něj, je totiž profi!.")



class PlantCodeFinder(tk.Frame):
    def __init__(self, master=None):
        super().__init__(master)
        self.master = master
        self.master.title("Olinky aplikace na pasy :)")
        self.added_codes = []
        self.grid()
        self.create_widgets()

        self.body = '''
        <html>
        <head>
        <style>
            body {
                font-family: Arial, sans-serif;
                font-size: 12px;
            }
            .larger-text {
                font-size: 18px;
            }
        </style>
        </head>
        <body>
            <b>Tento email je generován automaticky, a proto vás žádáme, abyste na něj neodpovídali. Případné dotazy nebo zprávy prosím zasílejte na adresu obchod@tropik.cz</b>.
        <br>
        <br>    
            Vážený zákazníku,
        <br>
        <br>
        ode dne 14. prosince 2019 nabývá účinnosti nové nařízení Evropského parlamentu, a to nám ukládá jako prodejci povinnosti při prodeji rostlin dodat zákazníkovi Rostlinolékařský pas. Je to z důvodu, aby se evidoval pohyb prodávaných rostlin po území Evropské unie.
        <br> 
        <br>
        <b>Více informací, proč došlo k této povinnosti se můžete dočíst na našem webu, v horním panelu (Nákupy, registrace). Vám, jakožto zákazníkovi, z toho neplyne žádná povinnost a na tuto automaticky generovanou zprávu neodpovídejte.</b> 
        <br>
        <br>
        <b class="larger-text">Více informací o nařízení</b><br>
        <br>
        Nařízení Evropského parlamentu a Rady (EU) 2016/2031 o ochranných opatřeních proti škodlivým organismům rostlin (dále jen „nařízení“). Dle čl. 65 tohoto nařízení je pro internetové prodejce rostlin, rostlinných produktů a jiných předmětů, podléhajících fytosanitární regulaci (dále jen regulované komodity), stanovena povinnost registrace pro rostlinolékařské účely, a to bez výjimky. Dále je dle čl. 79 a čl. 81 nařízení stanovena povinnost opatřovat regulované komodity při internetovém obchodování (smlouvy uzavřené na dálku) rostlinolékařským pasem, a to i v případě dodávek těchto komodit přímo konečným uživatelům. 
        <br>
        <br>
        Veškeré informace ohledně zákazu dovozu určitých rostlin, zvláštních a rovnocenných požadavcích, které musí při dovozu na území EU nebo při přemísťování na tomto území, vysoce rizikových rostlinách, rostlinných produktech či jiných předmětech, výjimkách z požadavku na rostlinolékařské osvědčení pro malá množství určitých rostlin naleznete na stránkách ÚKZÚZ http://eagri.cz/public/web/ukzuz/portal/ <br>
        <br>
        <br>
        Rostlinolékařský pas naleznete v příloze. <br>
        </body>
        </html>
        ''' 
        self.attachments_folder = os.path.join(os.getcwd(), 'pdf')
    def create_widgets(self):
        # Main frame
        main_frame = ttk.Frame(self, padding=20)
        main_frame.grid(row=0, column=0)

        # Configure the grid to have a minimum size and to expand with the window
        for i in range(4):  # Increase the range to 4
            main_frame.columnconfigure(i, weight=1, minsize=150)
            main_frame.rowconfigure(i, weight=1, minsize=50)

        # Section 1: Buttons
        section1_label = ttk.Label(main_frame, text="Hlavní funkce", font=("Helvetica", 12, "bold"))
        section1_label.grid(row=0, column=0, pady=(0, 10), sticky="w")

        buttons_frame = ttk.Frame(main_frame, borderwidth=2, relief="groove", padding=10)
        buttons_frame.grid(row=1, column=0, sticky="w")

        # Empty Label to create space
        ttk.Label(main_frame).grid(row=1, column=1)

        # Configure the grid to have a minimum size and to expand with the window
        for i in range(2):
            buttons_frame.columnconfigure(i, weight=1, minsize=150)
            buttons_frame.rowconfigure(i, weight=1, minsize=50)

        # Buttons grid
        self.process_button = ttk.Button(buttons_frame, text="Zpracovat faktury", command=self.process_pdfs)
        self.process_button.grid(row=0, column=0, pady=10, padx=10, sticky="nsew")

        self.manual_code_button = ttk.Button(buttons_frame, text="Připravit excely", command=self.create_excel)
        self.manual_code_button.grid(row=0, column=1, pady=10, padx=10, sticky="nsew")

        self.create_excel_button = ttk.Button(buttons_frame, text="Smazat vše", command=self.delete_codes_and_cz)
        self.create_excel_button.grid(row=3, column=0, pady=10, padx=10, sticky="nsew")

        self.process_pdf_button = ttk.Button(buttons_frame, text="Uložit Excely jako PDFka", command=self.save_all_excels_as_pdfs)
        self.process_pdf_button.grid(row=4, column=1, pady=10, padx=10, sticky="nsew")

        self.check_updates_button = ttk.Button(buttons_frame, text="Ořezat názvy", command=self.open_trim_names_window)
        self.check_updates_button.grid(row=1, column=1, pady=10, padx=10, sticky="nsew")

        self.save_excels_as_pdfs_button = ttk.Button(buttons_frame, text="Zpracovat soubory", command=self.process_files)
        self.save_excels_as_pdfs_button.grid(row=2, column=1, padx=10, pady=10, sticky="nsew")

        self.compress_button = ttk.Button(buttons_frame, text="Optimalizovat excely", command=self.compress_excel_files)
        self.compress_button.grid(row=2, column=0, padx=10, pady=10, sticky="nsew")

        self.process_plants_button = ttk.Button(buttons_frame, text="Získat rostliny", command=self.process_txts_for_plants)
        self.process_plants_button.grid(row=1, column=0, pady=10, padx=10, sticky="nsew")

        self.remove_duplicates_button = ttk.Button(buttons_frame, text="Odebrat duplicity a listy", command=self.remove_duplicates)
        self.remove_duplicates_button.grid(row=3, column=0, padx=10, pady=10, sticky="nsew")

        self.insert_image_button = ttk.Button(buttons_frame, text="Vložit EU obrázky", command=self.insert_image_to_excel)
        self.insert_image_button.grid(row=4, column=0, pady=10, padx=10, sticky="nsew")

        self.optimize_button = ttk.Button(buttons_frame, text="Vyřešit ENDOPA", command=self.optimize_excel_files)
        self.optimize_button.grid(row=3, column=1, padx=10, pady=10, sticky="nsew")

        # Label verze
        self.version_label = ttk.Label(main_frame, text=current_version, font=("Helvetica", 10, "bold"))
        self.version_label.grid(row=0, column=2, pady=10, sticky="e")

        # Section 2: Checkboxes
        section2_label = ttk.Label(main_frame, text="Nastavení", font=("Helvetica", 12, "bold"))
        section2_label.grid(row=4, column=0, pady=(20, 10), sticky="w")

        checkboxes_frame = ttk.Frame(main_frame, padding=10)
        checkboxes_frame.grid(row=5, column=0, sticky="w")

        self.delete_checkbox_var = tk.BooleanVar()
        self.delete_checkbox = ttk.Checkbutton(checkboxes_frame, text="Smazat soubory, kde je vše v poho", variable=self.delete_checkbox_var)
        self.delete_checkbox.grid(row=0, column=0, padx=10, pady=10, sticky="w")

        self.fulfill_cz_checkbox_var = tk.BooleanVar()
        self.fulfill_cz_checkbox = ttk.Checkbutton(checkboxes_frame, text="Vyplnit CZ ke kodum", variable=self.fulfill_cz_checkbox_var)
        self.fulfill_cz_checkbox.grid(row=1, column=0, padx=10, pady=10, sticky="w")

        # Section 3: Output console
        self.output_console = tk.Text(main_frame, wrap=tk.WORD, height=15, width=100, relief="sunken", borderwidth=1)  # Increase height
        self.output_console.grid(row=7, column=0, columnspan=4, padx=10, pady=10)


        # Quit button
        self.quit_button = ttk.Button(main_frame, text="Ukončit", command=self.quit)
        self.quit_button.grid(row=9, column=3, padx=10, pady=10, sticky="e")
                
        # Section 4: New Buttons
        section4_label = ttk.Label(main_frame, text="Doplňující funkce", font=("Helvetica", 12, "bold"))
        section4_label.grid(row=0, column=2, pady=(0, 10), sticky="w")  # Change column to 2

        new_buttons_frame = ttk.Frame(main_frame, borderwidth=2, relief="groove", padding=10)
        new_buttons_frame.grid(row=1, column=2, sticky="w")  # Change column to 2

        # Configure the grid to have a minimum size and to expand with the window
        #for i in range(2):
         #   new_buttons_frame.columnconfigure(i, weight=1, minsize=150)
          #  new_buttons_frame.rowconfigure(i, weight=1, minsize=50)

        # New buttons grid

        self.delete_codes_button = ttk.Button(new_buttons_frame, text="Přidat kód", command=self.manually_add_code)
        self.delete_codes_button.grid(row=0, column=0, padx=10, pady=10, sticky="nsew")

        self.manage_added_codes_button = ttk.Button(new_buttons_frame, text="Spravovat kódy v paměti", command=self.manage_added_codes)
        self.manage_added_codes_button.grid(row=0, column=1, padx=10, pady=10, sticky="nsew")

        self.load_temporary_button = ttk.Button(new_buttons_frame, text="Načíst z dočasné", command=self.load_from_temporary)
        self.load_temporary_button.grid(row=1, column=0, padx=10, pady=10, sticky="nsew")

        self.show_unmatched_button = ttk.Button(new_buttons_frame, text="Zobrazit neshody", command=self.show_unmatched_names)
        self.show_unmatched_button.grid(row=1, column=1, pady=10, padx=10, sticky="nsew")

        self.send_email_button = ttk.Button(new_buttons_frame, text="Pochvala pro Martínka", command=self.send_email)
        self.send_email_button.grid(row=2, column=0, padx=10, pady=10, sticky="nsew")
        

        self.send_second_email_button = ttk.Button(new_buttons_frame, text="Bída pro Martínka", command=self.send_second_email)
        self.send_second_email_button.grid(row=2, column=1, padx=10, pady=10, sticky="nsew")


        self.send_emails_button = ttk.Button(new_buttons_frame, text="Odeslat e-maily blbečkům", command=self.manage_recipients)
        self.send_emails_button.grid(row=3, column=0, pady=10, padx=10, sticky="w")

        self.create_excel_button = ttk.Button(new_buttons_frame, text="Smazat všecko v Excelu", command=self.delete_codes_and_cz)
        self.create_excel_button.grid(row=3, column=1, pady=10, padx=10, sticky="w")

        self.trim_names_button = ttk.Button(main_frame, text="Zkontrolovat aktualizace", command=self.check_for_updates)
        self.trim_names_button.grid(row=9, column=0, pady=10, padx=10, sticky="w")

        self.open_tropik_button = ttk.Button(new_buttons_frame, text="Tropik.cz", command=self.open_tropik)
        self.open_tropik_button.grid(row=4, column=0, pady=10, padx=10, sticky="nsew")

        self.open_eshop_tropik_button = ttk.Button(new_buttons_frame, text="Eshop Tropik", command=self.open_eshop_tropik)
        self.open_eshop_tropik_button.grid(row=4, column=1, pady=10, padx=10, sticky="nsew")

        # Initialize instance variables
        self.template_wb = openpyxl.load_workbook(self.get_template())
        self.codes = {}
        for row in range(2, self.template_wb.active.max_row + 1):
            name = self.template_wb.active.cell(row=row, column=1).value
            code = self.template_wb.active.cell(row=row, column=2).value
            self.codes[name] = code

    def check_for_updates(self):
        url = 'https://api.github.com/repos/{owner}/{repo}/releases/latest'
        try:
            response = requests.get(url.format(owner='shaneomac1337', repo='rostlinolekarske_pasy'))

            if response.status_code == requests.codes.ok:
                latest_release = response.json()
                latest_version = latest_release['tag_name'][1:]
                if semver.compare(current_version[1:], latest_version) < 0:
                    # Display a message box to the user
                    result = messagebox.askyesno('Aktualizace dostupná', 'Nová verze toolu na pasy je k dispozici, přeje si Olinka stáhnout novou verzi z webu?')

                    if result:
                        # Open the Github page for the latest release in the user's default web browser
                        url = latest_release['html_url']
                        webbrowser.open_new(url)

                        # An update is available, download the asset(s) that match your platform and architecture
                        assets = latest_release['assets']
                        for asset in assets:
                            if 'Windows' in asset['name'] and 'x86_64' in asset['name']:
                                download_url = asset['browser_download_url']
                                r = requests.get(download_url)
                                # Save the downloaded asset to a file
                                with open('aplikace_v.0.4.0.exe', 'wb') as f:
                                    f.write(r.content)
                    else:
                        # Do nothing if the user clicks "No"
                        pass
                else:
                    # No update available
                    messagebox.showinfo('Žádné aktualizace', 'Pro Olinku není k dispozici bohužel žádná aktualizace')
            else:
                # Failed to retrieve latest release info
                messagebox.showerror('Chyba', 'Nedokázal jsem zjistit informace o nejnovější verzi z GitHubu.')
        except requests.exceptions.RequestException:
            # Display an error message box
            messagebox.showerror("Chyba", "Olinka nemá internet a chtěla by novou verzi, smůla ty naivko!! Hezky pojedeš na starý.")


    def has_excel_files(self):
        for filename in os.listdir('.'):
            if filename.endswith('.xlsx') and filename != 'template.xlsx':
                return True
        return False

    def get_template(self):
        template_path = 'template.xlsx'
        return template_path

    def load_custom_template(self):
        file_path = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx")])
        if file_path:
            self.template_wb = openpyxl.load_workbook(file_path)
            self.codes = {}
            for row in range(2, self.template_wb.active.max_row + 1):
                name = self.template_wb.active.cell(row=row, column=1).value
                code = self.template_wb.active.cell(row=row, column=2).value
                self.codes[name] = code
            self.output_console.insert(tk.END, f"Olinka podstrčila tuhle šablonu: {file_path}\n")
            self.output_console.see(tk.END)  # Auto-scroll to the end
            self.output_console.update()  # Ensure the output console is updated

    def unload_custom_template(self):
        self.codes = {}
        self.template_wb = openpyxl.load_workbook(self.get_template())
        for row in range(2, self.template_wb.active.max_row + 1):
            name = self.template_wb.active.cell(row=row, column=1).value
            code = self.template_wb.active.cell(row=row, column=2).value
            self.codes[name] = code
        self.output_console.insert(tk.END, "Defaultní šablona načtena.\n")
        self.output_console.see(tk.END)  # Auto-scroll to the end
        self.output_console.update()  # Ensure the output console is updated   

    def show_unmatched_names(self):
        unmatched_names = []
        for filename in os.listdir('.'):
            if not filename.endswith('.xlsx') or filename == 'template.xlsx' or filename == 'temporary.xlsx':
                continue

            wb = openpyxl.load_workbook(filename)
            for sheet in wb:
                ws = wb[sheet.title]
                for row in range(13, ws.max_row + 1):
                    name = ws.cell(row=row, column=3).value
                    if name:
                        closest_match = max(self.codes.keys(), key=lambda x: fuzz.ratio(x, name))
                        if fuzz.ratio(closest_match, name) < 80:
                            unmatched_names.append((filename, sheet.title, name, row))

        if unmatched_names:
            message = "Nenalezené shody:\n\n"
            for filename, sheet_title, name, row in unmatched_names:
                message += f"{filename} - {sheet_title} - {name} (řádek {row})\n"
            self.output_console.insert(tk.END, message)
            self.output_console.see(tk.END)  # Auto-scroll to the end
            self.output_console.update()  # Ensure the output console is updated
        else:
            self.output_console.insert(tk.END, "Nenašel jsem neshody.\n")
            self.output_console.see(tk.END)  # Auto-scroll to the end
            self.output_console.update()  # Ensure the output console is updated


    def process_files(self):
        # Check if there are any Excel files to process
        if not self.has_excel_files():
            messagebox.showwarning("Kde jsou sakra Excely???", "Nejdřív mi musí Olinka navalit nějaký ten Excel do stejné složky (muže jich být nespočet), kde jsem byl spuštěn. Tak je koukej vysolit!!! Nebo vyšli Edíka s Kačenkou ať je donesou...")
            return

        # Create a "missing" folder if it doesn't exist
        if not os.path.exists("missing"):
            os.makedirs("missing")

        # Process all Excel files in the current directory
        for filename in os.listdir('.'):
            # Skip non-Excel files and the template file
            if not filename.endswith('.xlsx') or filename == 'template.xlsx' or filename == 'temporary.xlsx':
                continue

            # Load the Excel file
            wb = openpyxl.load_workbook(filename)

            # Process all sheets in the Excel file
            for sheet in wb:
                # Get the active sheet
                ws = wb[sheet.title]

                # Set the width of Column C to 48.71
                ws.column_dimensions['C'].width = 48.71

                # Process all rows in the sheet and add missing plant codes
                missing_names = []
                for row in range(13, ws.max_row + 1):
                    name = ws.cell(row=row, column=3).value
                    if name:
                        # Find the closest matching plant code in the dictionary
                        closest_match = max(self.codes.keys(), key=lambda x: fuzz.ratio(x, name))
                        if fuzz.ratio(closest_match, name) >= 80:
                            ws.cell(row=row, column=4).value = self.codes[closest_match]
                            if self.fulfill_cz_checkbox_var.get():
                                ws.cell(row=row, column=5).value = 'CZ'
                        else:
                            if self.fulfill_cz_checkbox_var.get():
                                ws.cell(row=row, column=5).value = 'CZ'
                            missing_names.append((name, row))                        

                # Save the modified Excel file
                wb.save(filename)

                # Write missing plant names to a text file in the "missing" folder
                if missing_names:
                    with open(f"missing/{filename}_{sheet.title}_missing_names.txt", 'w', encoding='utf-8') as f:
                        f.write(f"Missing names v excel sheetu:  '{sheet.title}' v souboru '{filename}':\n")
                        for name, row in missing_names:
                            f.write(f"Jméno '{name}' chybí na řádce {row}\n")
                elif not self.delete_checkbox_var.get():
                    with open(f"missing/{filename}_{sheet.title}_is_okay.txt", 'w', encoding='utf-8') as f:
                        f.write(f"Všechna jména se nachází v excel sheetu: '{sheet.title}' v souboru: '{filename}'\n")

        # Delete files without missing plant names if the checkbox is checked
        if self.delete_checkbox_var.get():
            files_to_delete = [filename for filename in os.listdir("missing") if filename.endswith("_is_okay.txt")]
            for filename in files_to_delete:
                os.remove(f"missing/{filename}")
                self.output_console.insert(tk.END, f"Deleted file: {filename}\n")
                self.output_console.see(tk.END)  # Auto-scroll to the end
                self.output_console.update()  # Ensure the output console is updated

        self.output_console.insert(tk.END, "Hotovo Olinko!\n")
        self.output_console.see(tk.END)  # Auto-scroll to the end
        self.output_console.update()  # Ensure the output console is updated

        messagebox.showinfo("Povedlo se!", "Teď už to jen stačí poslat blbečkům na mail :).")

    def delete_codes_and_cz(self):
        if not self.has_excel_files():
            messagebox.showerror("Chyba", "Olinka už mi zase nedala Excely, dělá si ze mě prdel??")
        else:
            for filename in os.listdir('.'):
                if not filename.endswith('.xlsx') or filename in ['template.xlsx', 'temporary.xlsx']:
                    continue 
                wb = openpyxl.load_workbook(filename)
                for sheet in wb:
                    ws = wb[sheet.title]
                    for row in range(13, ws.max_row + 1):
                        ws.cell(row=row, column=3).value = None  # This line deletes the text from column C
                        ws.cell(row=row, column=4).value = None
                        ws.cell(row=row, column=5).value = None
                wb.save(filename)

            self.output_console.insert(tk.END, "Je to fuč..\n")
            self.output_console.see(tk.END)  # Auto-scroll to the end
            self.output_console.update()  # Ensure the output console is updated

    def create_missing_folder(self):
        if not os.path.exists("missing"):
            os.makedirs("missing")
            self.output_console.insert(tk.END, "Složka missing vytvořena.\n")
            self.output_console.see(tk.END)  # Auto-scroll to the end
            self.output_console.update()  # Ensure the output console is updated
        else:
            self.output_console.insert(tk.END, "Složka missing už existuje.\n")
            self.output_console.see(tk.END)  # Auto-scroll to the end
            self.output_console.update()  # Ensure the output console is updated
    def save_excel_as_pdf(self, excel_file, sheet_name, queue, counter):
        """Save the given sheet in excel file as pdf and merge it with original invoice."""
        pdf_folder = "pdf"
        if not os.path.exists(pdf_folder):
            os.makedirs(pdf_folder)

        pdf_file = f"{pdf_folder}/{sheet_name}.pdf"
    
        wb = None
        try:
            pythoncom.CoInitialize()
            xlApp = win32.Dispatch("Excel.Application")
            xlApp.Visible = False

            wb = xlApp.Workbooks.Open(os.path.abspath(excel_file), ReadOnly=1)
            ws = wb.Worksheets(sheet_name)
            ws.PageSetup.Orientation = 2  # 2 represents landscape orientation
            ws.PageSetup.Zoom = False  # Turn off Zoom property
            ws.PageSetup.FitToPagesWide = 1  # Fit to 1 page wide
            ws.PageSetup.FitToPagesTall = 1  # Fit to 1 page tall
            ws.ExportAsFixedFormat(0, os.path.abspath(pdf_file))

            # After creating the PDF from Excel, merge it with the original invoice
            invoice_pdf = f"faktury/single_invoices/{sheet_name}.pdf"  # Updated path to look in single_invoices
            if os.path.exists(invoice_pdf):
                merger = PdfMerger()
                
                # Add the plant passport PDF (generated from Excel)
                merger.append(pdf_file)
                
                # Add the original invoice PDF
                merger.append(invoice_pdf)
                
                # Save the merged PDF, temporarily rename the original
                temp_file = f"{pdf_folder}/temp_{sheet_name}.pdf"
                merger.write(temp_file)
                merger.close()
                
                # Replace the original PDF with the merged version
                os.remove(pdf_file)
                os.rename(temp_file, pdf_file)
                
                queue.put(f"Uloženo jako: {sheet_name} (včetně faktury)\n")
            else:
                queue.put(f"Uloženo jako: {sheet_name} (bez faktury - nenalezena)\n")

        except Exception as e:
            queue.put(f"Chyba při zpracování {sheet_name}: {str(e)}\n")

        finally:
            if wb is not None:
                wb.Close(SaveChanges=False)
            xlApp.Quit()

        # Increment the counter
        counter[0] += 1

    def save_all_excels_as_pdfs(self):
        has_excel_files = False
        q = queue.Queue()
        counter = [0]  # Use a list so that the value is mutable inside the nested function

        threads = []
        for filename in os.listdir('.'):
            if filename.endswith('.xlsx') and filename != 'template.xlsx' and filename != 'temporary.xlsx':
                has_excel_files = True
                wb = openpyxl.load_workbook(filename)
                for sheet in wb:
                    t = threading.Thread(target=self.save_excel_as_pdf, args=(filename, sheet.title, q, counter))
                    t.start()
                    threads.append(t)

        # Function to update the GUI
        def update_gui():
            while not q.empty():
                message = q.get()
                self.output_console.insert(tk.END, message)
                self.output_console.see(tk.END)  # Auto-scroll to the end
                self.output_console.update()  # Ensure the output console is updated

            # If all threads have finished, stop updating the GUI
            if counter[0] == len(threads):
                if not has_excel_files:
                    messagebox.showerror("Chyba", "A teď zase Olinka nedodala Excely na konvertování do PDF. Bože muj.")
                else:
                    self.output_console.insert(tk.END, "Všechny listy v excelu byly uloženy jako samostatné PDF.\n")
                    self.output_console.see(tk.END)  # Auto-scroll to the end
                    self.output_console.update()  # Ensure the output console is updated
            else:
                self.output_console.after(100, update_gui)  # Schedule the next call

        # Start the GUI update
        self.output_console.after(100, update_gui)

    def insert_image_to_excel(self):
        def insert_image(excel_file_path, image_file_path, cell_name, row_height=None, column_width=None, pic_width=None, pic_height=None):
            # Open Excel
            excel = win32.gencache.EnsureDispatch('Excel.Application')
            excel.Visible = False  # If you want Excel Application to be visible during execution, set this to True

            # Open Workbook
            wb = excel.Workbooks.Open(excel_file_path)

            # Iterate through each sheet in the workbook
            for sheet in wb.Sheets:

                # Get cell dimensions
                target_cell = sheet.Range(cell_name)
                top = target_cell.Top
                left = target_cell.Left

                # Set row height and column width for the entire sheet if provided
                if row_height is not None:
                    sheet.Rows.RowHeight = row_height
                if column_width is not None:
                    sheet.Columns.ColumnWidth = column_width

                # Add picture
                pic = sheet.Pictures().Insert(image_file_path)

                # Set image position
                pic.Top = top
                pic.Left = left

                # Set image size
                if pic_width is not None:
                    pic.Width = pic_width
                if pic_height is not None:
                    pic.Height = pic_height

            # Save and Close
            wb.Save()
            wb.Close()

            # Quit Excel
            excel.Quit()

        # Get the current working directory
        directory_path = os.getcwd()
        cell_name = "C6"

        # Set your desired row height
        row_height = 15.75  # Adjust this value to change the row height for all rows in the sheet

        # Set your desired picture width and height
        pic_width = 147  # Adjust this value to set the picture width
        pic_height = 80  # Adjust this value to set the picture height

        # Get a list of all Excel files in the directory
        excel_files = [f for f in os.listdir(directory_path) if f.endswith('.xlsx') or f.endswith('.xls')]

        # Run the function on each Excel file
        for excel_file in excel_files:
            if excel_file == 'template.xlsx' or excel_file == 'temporary.xlsx':
                continue  # Skip the template file and temporary file
            excel_file_path = os.path.join(directory_path, excel_file)
            image_file_path = os.path.join(directory_path, "eu.png")  # The image file is in the same directory as the Excel files
            insert_image(excel_file_path, image_file_path, cell_name, row_height, None, pic_width, pic_height)
            self.output_console.insert(tk.END, f"Obrázek vložen do: {excel_file}\n")
            self.output_console.see(tk.END)  # Auto-scroll to the end
            self.output_console.update()  # Ensure the output console is updated      

    def manually_add_code(self):
        def populate_listbox():
            # Clear the listbox
            name_listbox.delete(0, tk.END)

            # Get the unmatched names
            unmatched_names = get_unmatched_names()

            # Create a set to store the names that have been added to the listbox
            added_names = set()

            # Populate the listbox
            for _, _, name, _ in unmatched_names:
                # Only add the name to the listbox if it hasn't been added before
                if name not in added_names:
                    name_listbox.insert(tk.END, name)
                    added_names.add(name)  # Add the name to the set of added names

        def submit_code():
            selected_index = name_listbox.curselection()[0]  # Get the index of the selected item
            selected_name = name_listbox.get(selected_index)
            code = code_entry.get()
            if selected_name and code:
                self.codes[selected_name] = code

                # Add the new code to the list of added codes
                self.added_codes.append((selected_name, code, "CZ"))
                populate_listbox()

                # Clear the code entry field for the next input
                code_entry.delete(0, 'end')

                # Always select the first item in the listbox after refresh
                if name_listbox.size() > 0:  # Check if the listbox is not empty
                    name_listbox.selection_clear(0, tk.END)  # Clear all selections
                    name_listbox.selection_set(0)  # Select the first item
                    name_listbox.see(0)  # Make sure the first item is visible
                    name_listbox.event_generate("<<ListboxSelect>>")  # Generate a ListboxSelect event to update the code_entry
                    code_entry.focus_set()  # Set the focus to the code_entry field

                self.output_console.insert(tk.END, f"Ručně přidán kód: {selected_name} - {code}\n")
                self.output_console.see(tk.END)
                self.output_console.update()

        def write_to_temporary():
            # Check if there are any added codes
            if not self.added_codes:
                self.output_console.insert(tk.END, "Není tu nic, co by Olinka mohla zapsat do dočasné :(\n")
                self.output_console.see(tk.END)
                self.output_console.update()
                return

            # Load the temporary workbook
            temporary_wb = openpyxl.load_workbook('temporary.xlsx')
            temporary_ws = temporary_wb.active

            # Create a border
            border = openpyxl.styles.Border(left=openpyxl.styles.Side(style='thin'),
                                            right=openpyxl.styles.Side(style='thin'),
                                            top=openpyxl.styles.Side(style='thin'),
                                            bottom=openpyxl.styles.Side(style='thin'))

            # Write the added codes to the temporary workbook
            for selected_name, code, cz in self.added_codes:
                # Find the correct row to insert the new plant name and code
                insert_row = None
                for row in range(2, temporary_ws.max_row + 1):
                    existing_code = temporary_ws.cell(row=row, column=2).value  # assuming codes are in column 2
                    if existing_code:
                        # Extract the numerical part from the existing code and the new code
                        existing_code_num = int(''.join(filter(str.isdigit, existing_code)))
                        new_code_num = int(''.join(filter(str.isdigit, code)))

                        if new_code_num < existing_code_num:
                            insert_row = row
                            break

                if insert_row is None:
                    # If the new plant name is greater than all existing names, append it to the end
                    insert_row = temporary_ws.max_row + 1

                # Insert a new row at the correct position
                temporary_ws.insert_rows(insert_row)

                # Write the new plant name and code to the temporary
                name_cell = temporary_ws.cell(row=insert_row, column=1)
                code_cell = temporary_ws.cell(row=insert_row, column=2)
                cz_cell = temporary_ws.cell(row=insert_row, column=3)

                name_cell.value = selected_name
                code_cell.value = code
                cz_cell.value = cz  # add 'CZ' to the column next to the code

                # Or, if you want to add it to the output console in your GUI:
                self.output_console.insert(tk.END, f"Zapsáno do dočasné: {selected_name} - {code} - {cz}\n")
                self.output_console.see(tk.END)
                self.output_console.update()

                # Apply the border to the cells
                name_cell.border = border
                code_cell.border = border
                cz_cell.border = border  # apply the border to the 'CZ' cell

            # Save the temporary workbook
            temporary_wb.save('temporary.xlsx')


        def clear_temporary():
            # Load the temporary workbook
            temporary_wb = openpyxl.load_workbook('temporary.xlsx')
            temporary_ws = temporary_wb.active

            # Get the number of rows
            num_rows = temporary_ws.max_row

            # Check if the workbook is empty
            if num_rows == 1 and not any(temporary_ws.cell(row=1, column=i).value for i in range(1, temporary_ws.max_column + 1)):
                self.output_console.insert(tk.END, "V dočasné nic není.\n")
                self.output_console.see(tk.END)
                self.output_console.update()
                return

            # Log and delete all rows
            for row in range(1, num_rows + 1):
                name = temporary_ws.cell(row=row, column=1).value
                code = temporary_ws.cell(row=row, column=2).value
                cz = temporary_ws.cell(row=row, column=3).value

                # Or, if you want to add it to the output console in your GUI:
                self.output_console.insert(tk.END, f"Smazáno z dočasné: {name} - {code} - {cz}\n")
                self.output_console.see(tk.END)
                self.output_console.update()

            # Delete all rows
            temporary_ws.delete_rows(1, num_rows)

            # Save the temporary workbook
            temporary_wb.save('temporary.xlsx')
            

        def get_unmatched_names():
            unmatched_names = []
            added_names = set()  # Create a set to store the names that have been added
            for filename in os.listdir('.'):
                if not filename.endswith('.xlsx') or filename == 'template.xlsx' or filename == 'temporary.xlsx':
                    continue

                wb = openpyxl.load_workbook(filename)
                for sheet in wb:
                    ws = wb[sheet.title]
                    for row in range(13, ws.max_row + 1):
                        name = ws.cell(row=row, column=3).value
                        if name and name != "CZ" and name not in added_names:  # Check if the name hasn't been added before
                            closest_match = max(self.codes.keys(), key=lambda x: fuzz.ratio(x, name))
                            if fuzz.ratio(closest_match, name) < 80:
                                unmatched_names.append((filename, sheet.title, name, row))
                                added_names.add(name)  # Add the name to the set of added names
            return unmatched_names

        top = tk.Toplevel(self.master)
        top.title("Ručně přidat kód")

        # Set the default size of the window
        top.geometry("800x600")  # You can adjust the size as per your requirement


        # Configure the row and column containing the listbox to expand
        top.grid_rowconfigure(0, weight=1)
        top.grid_columnconfigure(1, weight=1)

        name_label = ttk.Label(top, text="Název:")
        name_label.grid(row=0, column=0, padx=10, pady=10, sticky="w")

        # Make the listbox expand when the window is resized
        name_listbox = tk.Listbox(top, selectmode=tk.SINGLE, exportselection=0)
        name_listbox.grid(row=0, column=1, padx=10, pady=10, sticky="nsew")  # "nsew" means the widget should expand in all directions

        unmatched_names = get_unmatched_names()
        for _, _, name, _ in unmatched_names:
            name_listbox.insert(tk.END, name)

        code_label = ttk.Label(top, text="Kód:")
        code_label.grid(row=1, column=0, padx=10, pady=10, sticky="w")
        code_entry = ttk.Entry(top)
        code_entry.grid(row=1, column=1, padx=10, pady=10, sticky="w")

        submit_button = ttk.Button(top, text="Potvrdit", command=submit_code)
        submit_button.grid(row=2, column=1, padx=10, pady=10, sticky="e")

        # Add a "Write to Temporary" button to write the added codes to temporary.xlsx
        write_button = ttk.Button(top, text="Zapsat do dočasné", command=write_to_temporary)
        write_button.grid(row=3, column=1, padx=10, pady=10, sticky="e")

        # Add a "Clear Temporary" button to clear all data from temporary.xlsx
        clear_button = ttk.Button(top, text="Vyčistit dočasnou", command=clear_temporary)
        clear_button.grid(row=4, column=1, padx=10, pady=10, sticky="e")

        # Add a "Append to Template" button to append the data from temporary.xlsx to template.xlsx
        append_button = ttk.Button(top, text="Zapsat z dočasné do templaty", command=self.append_to_template)
        append_button.grid(row=5, column=1, padx=10, pady=10, sticky="e")


        def copy_to_clipboard(event):
            # Get the selected item
            selected_item = name_listbox.get(name_listbox.curselection())

            # Copy the selected item to the clipboard
            top.clipboard_clear()
            top.clipboard_append(selected_item)

        # Bind the Ctrl+C key to the copy_to_clipboard function
        name_listbox.bind('<Control-c>', copy_to_clipboard)

        def on_name_select(event):
            code_entry.focus_set()

        name_listbox.bind('<ButtonRelease-1>', on_name_select)
        # Bind the Enter key to the submit_code function
        top.bind('<Return>', lambda event: submit_code())


    def load_from_temporary(self):
        # Load the temporary workbook
        temporary_wb = openpyxl.load_workbook('temporary.xlsx')
        temporary_ws = temporary_wb.active

        # Get the number of rows
        num_rows = temporary_ws.max_row

        # Check if the workbook is empty
        if num_rows == 1 and not any(temporary_ws.cell(row=1, column=i).value for i in range(1, temporary_ws.max_column + 1)):
            self.output_console.insert(tk.END, "Dočasná je prázdná.\n")
            self.output_console.see(tk.END)
            self.output_console.update()
            return

        # Clear self.added_codes
        self.added_codes.clear()

        # Load the codes from the temporary workbook
        for row in range(2, num_rows + 1):
            name = temporary_ws.cell(row=row, column=1).value
            code = temporary_ws.cell(row=row, column=2).value
            cz = temporary_ws.cell(row=row, column=3).value

            # Add the code to self.codes and self.added_codes
            self.codes[name] = code
            self.added_codes.append((name, code, cz))

            # Print the loaded data to the console
            # Or, if you want to add it to the output console in your GUI:
            self.output_console.insert(tk.END, f"Načteno z dočasné: {name} - {code} - {cz}\n")
            self.output_console.see(tk.END)
            self.output_console.update()

    def append_to_template(self):
        # Open Excel
        excel = win32.gencache.EnsureDispatch('Excel.Application')

        # Open the workbooks
        temp_wb = excel.Workbooks.Open(os.path.abspath('temporary.xlsx'))
        template_wb = excel.Workbooks.Open(os.path.abspath('template.xlsx'))

        # Get the worksheets
        temp_ws = temp_wb.Worksheets(1)
        template_ws = template_wb.Worksheets(1)

        # Get the last row in the template workbook
        last_row = template_ws.Cells(template_ws.Rows.Count, 1).End(win32.constants.xlUp).Row

        # Get the last row in the temporary workbook
        last_row_temp = temp_ws.Cells(temp_ws.Rows.Count, 1).End(win32.constants.xlUp).Row

        # Append the data from temporary.xlsx to the end of template.xlsx
        range_to_copy = temp_ws.Range(f"A2:C{last_row_temp}")
        range_to_copy.Copy(template_ws.Range(f"A{last_row + 1}"))

        # Print each line to the output console
        for i in range(2, last_row_temp + 1):
            name = temp_ws.Cells(i, 1).Value
            code = temp_ws.Cells(i, 2).Value
            cz = temp_ws.Cells(i, 3).Value
            self.output_console.insert(tk.END, f"Zapsáno: {name} - {code} - {cz}\n")
            self.output_console.see(tk.END)
            self.output_console.update()

        # Save and close the workbooks
        template_wb.Save()
        temp_wb.Close(SaveChanges=False)
        template_wb.Close()

        # Quit Excel
        excel.Quit()

        # Print a message to the output console
        self.output_console.insert(tk.END, "Data z dočasné byly zapsány do template.xlsx.\n")
        self.output_console.see(tk.END)
        self.output_console.update()

    def forget_added_codes(self, print_message=True):
        # Check if there are any added codes
        if not self.added_codes:
            self.output_console.insert(tk.END, "Žádné kódy k zapomenutí.\n")
            self.output_console.see(tk.END)
            self.output_console.update()
            return

        # Remove the added codes from self.codes
        for name, code, cz in self.added_codes:
            if name in self.codes:
                del self.codes[name]

            # Print a message to the console for each code that is removed
            self.output_console.insert(tk.END, f"Odebrán kód: {name} - {code} - {cz}\n")
            self.output_console.see(tk.END)
            self.output_console.update()

        # Clear the added_codes list
        self.added_codes.clear()

        # Print a message to the console
        if print_message:
            self.output_console.insert(tk.END, "Všechny ručně přidané kódy byly zapomenuty.\n")
            self.output_console.see(tk.END)
            self.output_console.update()

    def manage_added_codes(self):
        # Create a new dialog
        dialog = tk.Toplevel(self)  # Use 'self' to refer to the main window
        dialog.title("Spravovat kody v paměti")
        dialog.geometry("800x600")  # Set the size of the dialog

        # Create a frame to contain the listbox and scrollbar
        frame = tk.Frame(dialog)
        frame.pack(fill="both", expand=True, padx=10, pady=10)

        # Create a scrollbar
        scrollbar = tk.Scrollbar(frame)
        scrollbar.pack(side="right", fill="y")

        # Create a listbox to display the added codes
        listbox = tk.Listbox(frame, yscrollcommand=scrollbar.set)
        listbox.pack(side="left", fill="both", expand=True)

        # Configure the scrollbar to scroll the listbox
        scrollbar.config(command=listbox.yview)

        # Populate the listbox with the added codes
        for name, code, cz in self.added_codes:
            listbox.insert(tk.END, f"{name} - {code} - {cz}")

        # Create a function to delete the selected code
        def delete_selected_code():
            selected_index = listbox.curselection()[0]  # Get the index of the selected item
            selected_code = self.added_codes[selected_index]

            # Remove the selected code from self.codes and self.added_codes
            if selected_code[0] in self.codes:
                del self.codes[selected_code[0]]
            del self.added_codes[selected_index]

            # Print a message to the console
            self.output_console.insert(tk.END, f"Ručně přidaný kód {selected_code} byl zapomenut.\n")
            self.output_console.see(tk.END)
            self.output_console.update()

            # Remove the selected code from the listbox
            listbox.delete(selected_index)

        # Create a button to delete the selected code
        delete_button = ttk.Button(dialog, text="Zapomenout vybraný kód", command=delete_selected_code)
        delete_button.pack(padx=10, pady=(0, 10))

        # Create a function to delete all codes
        def delete_all_codes():
            self.forget_added_codes()  # Use the forget_added_codes function to delete all codes
            listbox.delete(0, tk.END)  # Clear the listbox

        # Create a button to delete all codes
        delete_all_button = ttk.Button(dialog, text="Zapomenout všechny kódy", command=delete_all_codes)
        delete_all_button.pack(padx=10, pady=(0, 10))

    def send_email(self):
        outlook = win32.Dispatch('outlook.application')
        mail = outlook.CreateItem(0)
        mail.Subject = 'Pochvala pro Martínka'
        mail.HTMLBody = '<p style="font-family:Roboto;">Milý Martínku,<br><br>moc ti děkuji za skvělou aplikaci na pasy, už bych si to dnes nedokázala vůbec představit bez ní, strašně moc mi pomáhá a můj život je díky tomu snadnější.<br><br>S pozdravem<br>Olinka</p>'
        mail.To = 'martinpenkava1@gmail.com'
        # Uncomment the line below if you want to send the email
        mail.Send()

        self.output_console.insert(tk.END, "Pochvala uspěšně odeslána Martínkovi na email, bude mít radost :)\n")
        self.output_console.see(tk.END)
        self.output_console.update()

    def send_second_email(self):
        outlook = win32.Dispatch('outlook.application')
        mail = outlook.CreateItem(0)
        mail.Subject = 'Nejde to vole'
        mail.HTMLBody = '<p style="font-family:Roboto;">Ahoj,<br><br>Bohužel, mám pro tebe špatné zprávy. Aplikace na pasy mi dnes nefunguje tak, jak bych si představovala Koukej s tim něco udělat a nebo ještě líp strč si ten svuj tool už do prdele, nikdo na něj neni zvědavej!<br><br>S pozdravem<br>Nasraná Olga</p>'
        mail.To = 'martinpenkava1@gmail.com'
        # Uncomment the line below if you want to send the email
        mail.Send()

        self.output_console.insert(tk.END, "Bída pro Martínka odeslána na email, Martínek bude smutný, zlá Olina!! :(\n")
        self.output_console.see(tk.END)
        self.output_console.update()

    def compress_excel_file(self, input_file):  # input_file should be the first argument after self
        # Load spreadsheet
        xl = pd.ExcelFile(input_file)

        # Load a sheet into a DataFrame by its name
        df_dict = {sheet_name: xl.parse(sheet_name) for sheet_name in xl.sheet_names}

        # Delete the ExcelFile object
        del xl


        # Create a new workbook
        wb = openpyxl.Workbook()

        for sheet_name, df in df_dict.items():
            # Create a new sheet
            ws = wb.create_sheet(title=sheet_name)

            # Write DataFrame to worksheet
            for r in dataframe_to_rows(df, index=False, header=True):
                ws.append(r)

            # Delete cells A1-E1
            for col in ['A', 'B', 'C', 'D', 'E']:
                ws[col + '1'].value = None

            # Merge cells D7-E7, D8-E8, and C6-C10
            ws.merge_cells('D7:E7')
            ws.merge_cells('D8:E8')
            ws.merge_cells('C6:C10')

            # Set font to bold and size to 14 for cells D7 and D8
            bold_font = Font(bold=True, size=14)
            ws['D7'].font = bold_font
            ws['D8'].font = bold_font

            # Set font to bold and size to 14 for cells D9 and D10 if they contain any text
            if ws['D9'].value:
                ws['D9'].font = bold_font
                ws.merge_cells('D9:E9')
            if ws['D10'].value:
                ws['D10'].font = bold_font
                ws.merge_cells('D10:E10') 

            # Set column widths
            ws.column_dimensions['B'].width = 4.71
            ws.column_dimensions['C'].width = 48.71
            ws.column_dimensions['D'].width = 21.14
            ws.column_dimensions['E'].width = 23.85
            ws.column_dimensions['F'].width = 4.28

            # Define border styles
            left_border = Border(left=Side(style='thin'))
            right_border = Border(right=Side(style='thin'))
            top_border = Border(top=Side(style='thin'))
            bottom_border = Border(bottom=Side(style='thin'))
            inside_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

            # Determine the last row with data in column C
            last_row = max(31, max((cell.row for cell in ws['C'] if cell.value is not None)) + 1)

            # Apply borders to the range B5:F31 (or B5:F32 if there's data in C31)
            for row in ws.iter_rows(min_row=5, max_row=last_row, min_col=2, max_col=6):
                for cell in row:
                    if cell.row == 5:
                        cell.border += top_border
                    if cell.row == last_row:
                        cell.border += bottom_border
                    if cell.column == 2:  # Column B
                        cell.border += left_border
                    if cell.column == 6:  # Column F
                        cell.border += right_border

            # Apply inside borders to the range C12:E30 (or further if there's data in C31 or beyond)
            for row in ws.iter_rows(min_row=12, max_row=last_row-1, min_col=3, max_col=5):
                for cell in row:
                    cell.border = inside_border

            # Make C12, D12, E12 align on center and bold
            center_aligned_bold_font = Font(bold=True)
            center_alignment = Alignment(horizontal='center')
            for col in ['C', 'D', 'E']:
                cell = ws[col + '12']
                cell.font = center_aligned_bold_font
                cell.alignment = center_alignment

            # Change the font of C13 and further down, D13 and further down, and E13 and further down to Arial 11
            arial_11_font = Font(name='Arial', size=10)
            for col in ['C', 'D', 'E']:
                for row in range(13, last_row):
                    cell = ws[col + str(row)]
                    cell.font = arial_11_font

            # Define a flag for whether "Endopa" is found and replaced
            endopa_replaced = False

            # Remove the word "Endopa" from cells D13 and further down
            for row in range(13, last_row):
                cell = ws['D' + str(row)]
                if cell.value and isinstance(cell.value, str) and "-CHZ-ENDOPA" in cell.value:
                    cell.value = cell.value.replace("-CHZ-ENDOPA", "")
                    endopa_replaced = True

           # If "Endopa" was found and replaced, set the text in cells D7, D8, D9, and D10
            if endopa_replaced:
                ws['D7'].value = "Rostlinolékařský pas"
                ws['D8'].value = "Plant Passport - PZ"
                ws['D9'].value = "ENDOPA"
                ws['D10'].value = "B: CZ - 0550"              

        # Delete the default sheet created and save the workbook
        del wb['Sheet']
        wb.save(input_file)

        # Force garbage collection
        gc.collect()

        # Display a message in the console
        self.output_console.insert(tk.END, f"Soubor {input_file} byl optimalizován.\n")
        self.output_console.see(tk.END)  # Auto-scroll to the end
        self.output_console.update()  # Ensure the output console is updated
    
    def compress_excel_files(self):
        excel_files = [f for f in os.listdir() if f.endswith('.xlsx') and f not in ['template.xlsx', 'temporary.xlsx','excluded_plant_names.xlsx']]
        for file in excel_files:
            self.compress_excel_file(input_file=file)  # Use 'self' to call the method

        # Display a message in the console when all files have been compressed
        self.output_console.insert(tk.END, "Všechny excely byly optimalizovány.\n")
        self.output_console.see(tk.END)  # Auto-scroll to the end
        self.output_console.update()  # Ensure the output console is updated


    def optimize_excel_file(self, input_file):  # input_file should be the first argument after self
        # Load spreadsheet
        xl = pd.ExcelFile(input_file)

        # Load a sheet into a DataFrame by its name
        df_dict = {sheet_name: xl.parse(sheet_name) for sheet_name in xl.sheet_names}

        # Delete the ExcelFile object
        del xl

        # Create a new workbook
        wb = openpyxl.Workbook()

        for sheet_name, df in df_dict.items():
            # Create a new sheet
            ws = wb.create_sheet(title=sheet_name)

            # Write DataFrame to worksheet
            for r in dataframe_to_rows(df, index=False, header=True):
                ws.append(r)

            # Delete cells A1-E1
            for col in ['A', 'B', 'C', 'D', 'E']:
                ws[col + '1'].value = None

            # Merge cells D7-E7, D8-E8, and C6-C10
            ws.merge_cells('D7:E7')
            ws.merge_cells('D8:E8')
            ws.merge_cells('C6:C10')

            # Set font to bold and size to 14 for cells D7 and D8
            bold_font = Font(bold=True, size=14)
            ws['D7'].font = bold_font
            ws['D8'].font = bold_font

            # Set font to bold and size to 14 for cells D9 and D10 if they contain any text
            if ws['D9'].value:
                ws['D9'].font = bold_font
                ws.merge_cells('D9:E9')
            if ws['D10'].value:
                ws['D10'].font = bold_font
                ws.merge_cells('D10:E10') 

            # Set column widths
            ws.column_dimensions['B'].width = 4.71
            ws.column_dimensions['C'].width = 48.71
            ws.column_dimensions['D'].width = 21.14
            ws.column_dimensions['E'].width = 23.85
            ws.column_dimensions['F'].width = 4.28

            # Define border styles
            left_border = Border(left=Side(style='thin'))
            right_border = Border(right=Side(style='thin'))
            top_border = Border(top=Side(style='thin'))
            bottom_border = Border(bottom=Side(style='thin'))
            inside_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

            # Determine the last row with data in column C
            last_row = max(31, max((cell.row for cell in ws['C'] if cell.value is not None)) + 1)

            # Apply borders to the range B5:F31 (or B5:F32 if there's data in C31)
            for row in ws.iter_rows(min_row=5, max_row=last_row, min_col=2, max_col=6):
                for cell in row:
                    if cell.row == 5:
                        cell.border += top_border
                    if cell.row == last_row:
                        cell.border += bottom_border
                    if cell.column == 2:  # Column B
                        cell.border += left_border
                    if cell.column == 6:  # Column F
                        cell.border += right_border

            # Apply inside borders to the range C12:E30 (or further if there's data in C31 or beyond)
            for row in ws.iter_rows(min_row=12, max_row=last_row-1, min_col=3, max_col=5):
                for cell in row:
                    cell.border = inside_border

            # Make C12, D12, E12 align on center and bold
            center_aligned_bold_font = Font(bold=True)
            center_alignment = Alignment(horizontal='center')
            for col in ['C', 'D', 'E']:
                cell = ws[col + '12']
                cell.font = center_aligned_bold_font
                cell.alignment = center_alignment

            # Change the font of C13 and further down, D13 and further down, and E13 and further down to Arial 11
            arial_11_font = Font(name='Arial', size=10)
            for col in ['C', 'D', 'E']:
                for row in range(13, last_row):
                    cell = ws[col + str(row)]
                    cell.font = arial_11_font

            # Define a flag for whether "Endopa" is found and replaced
            endopa_replaced = False

            # Remove the word "Endopa" from cells D13 and further down
            for row in range(13, last_row):
                cell = ws['D' + str(row)]
                if cell.value and isinstance(cell.value, str) and "-CHZ-ENDOPA" in cell.value:
                    # Print a message to the console before replacing "Endopa"
                    self.output_console.insert(tk.END, f"Řeším Endopu pro: {sheet_name} v souboru: {input_file}.\n")
                    self.output_console.see(tk.END)  # Auto-scroll to the end
                    self.output_console.update()  # Ensure the output console is updated

                    cell.value = cell.value.replace("-CHZ-ENDOPA", "")
                    endopa_replaced = True
        # If "Endopa" was found and replaced, set the text in cells D7, D8, D9, and D10
            if endopa_replaced:
                ws['D7'].value = "Rostlinolékařský pas"
                ws['D8'].value = "Plant Passport - PZ"
                ws['D9'].value = "ENDOPA"
                ws['D10'].value = "B: CZ - 0550"
           

        # Delete the default sheet created and save the workbook
        del wb['Sheet']
        wb.save(input_file)

        # Force garbage collection
        gc.collect()

        # Display a message in the console
        self.output_console.insert(tk.END, f"ENDOPA pro soubor {input_file} je v cajku.\n")
        self.output_console.see(tk.END)  # Auto-scroll to the end
        self.output_console.update()  # Ensure the output console is updated

    def optimize_excel_files(self):
        excel_files = [f for f in os.listdir() if f.endswith('.xlsx') and f not in ['template.xlsx', 'temporary.xlsx','excluded_plant_names.xlsx']]
        for file in excel_files:
            self.optimize_excel_file(input_file=file)  # Use 'self' to call the method

        # Display a message in the console when all files have been optimized
        self.output_console.insert(tk.END, "ENDOPA vyřešena.\n")
        self.output_console.see(tk.END)  # Auto-scroll to the end
        self.output_console.update()  # Ensure the output console is updated

      

    def manage_recipients(self):
        # Create a new dialog
        dialog = tk.Toplevel(self)  # Use 'self' to refer to the main window
        dialog.title("Spravovat adresáty a přílohy")
        dialog.geometry("800x600")  # Set the size of the dialog

        # Create a frame to contain the listbox and scrollbar
        frame = tk.Frame(dialog)
        frame.pack(fill="both", expand=True, padx=10, pady=10)

        # Create a scrollbar
        scrollbar = tk.Scrollbar(frame)
        scrollbar.pack(side="right", fill="y")

        # Create a listbox to display the recipients and attachments
        listbox = tk.Listbox(frame, yscrollcommand=scrollbar.set)
        listbox.pack(side="left", fill="both", expand=True)

        # Configure the scrollbar to scroll the listbox
        scrollbar.config(command=listbox.yview)

        # Get the directory of the current script
        script_dir = os.path.dirname(os.path.realpath(__file__))

        # Join the directory with the filename to get the full path of the file
        file_path = os.path.join(script_dir, 'mail_tool', 'recipients.xlsx')

        # Load the xlsx file
        email_list = pd.read_excel(file_path)

        # Create a label and entry for the BCC recipient
        bcc_label = tk.Label(dialog, text="Skrytou kopii poslat na:")
        bcc_label.pack(padx=10, pady=10)
        bcc_entry = tk.Entry(dialog, width=50)  # Increase the width value as needed
        bcc_entry.pack(padx=10, pady=10)
        bcc_entry.insert(0, "tropiktropik.cz@gmail.com")  # Set the default BCC recipient

        # Populate the listbox with the recipients and attachments
        for _, row in email_list.iterrows():
            recipient = row['Email']
            attachment_name = row['Attachment']
            listbox.insert(tk.END, f"{recipient} - {attachment_name}")


        # Create a function to send email to the selected recipient
        def send_email_to_selected_recipient():
            selected_index = listbox.curselection()[0]  # Get the index of the selected item
            selected_recipient = email_list.iloc[selected_index]

            recipient = selected_recipient['Email']
            attachment_name = selected_recipient['Attachment']
            bcc_recipient = bcc_entry.get()
            self.send_email_with_attachment(recipient, bcc_recipient, attachment_name)

        # Create a button to send email to the selected recipient
        send_button = tk.Button(dialog, text="Odeslat e-maily", command=send_email_to_selected_recipient)
        send_button.pack(padx=10, pady=10)

        # Create a function to send emails to all recipients
        def send_emails_to_all_recipients():
            # Get the BCC recipient from the entry field
            bcc_recipient = bcc_entry.get()

            # Open the log file in append mode
            with open('email.log', 'a') as f:
                for _, row in email_list.iterrows():
                    recipient = row['Email']
                    attachment_name = row['Attachment']
                    self.send_email_with_attachment(recipient, bcc_recipient, attachment_name)
                    
                    # Write the output to the log file
                    f.write(f"E-mail odeslán na {recipient} s přílohou {attachment_name}\n")

                    # Output to the console

                    dialog.update_idletasks()  # Force the GUI to update

            self.output_console.insert(tk.END, "Olinka poslala mail všem blbečkům\n")




        # Create a button to send emails to all recipients
        send_all_button = tk.Button(dialog, text="Odeslat e-maily všem", command=send_emails_to_all_recipients)
        send_all_button.pack(padx=10, pady=10)

    
    def send_email_with_attachment(self, recipient, bcc_recipient, attachment_name):
        self.output_console.insert(tk.END, f"Připravuji e-mail pro: {recipient} s přílohou: {attachment_name}\n")
        self.output_console.update_idletasks()  # Force the GUI to update  # Force the GUI to update    
        outlook = win32.Dispatch('outlook.application')
        mail = outlook.CreateItem(0)
        attachment_name_without_ext, _ = os.path.splitext(attachment_name)  # Remove the .pdf suffix
        mail.Subject = f'Rostlinolékařský pas k vaší faktuře č. {attachment_name_without_ext}'  # Format the subject with the attachment name without the .pdf suffix
        mail.HTMLBody = self.body
        mail.To = recipient
        mail.BCC = bcc_recipient

        if not pd.isna(attachment_name):
            attachment_path = os.path.join(self.attachments_folder, attachment_name)
            if os.path.isfile(attachment_path):
                mail.Attachments.Add(attachment_path)
            else:
                self.output_console.insert(tk.END, f"Příloha nenalezena: {attachment_path}. Neodeslal jsem e-mail pro {recipient}.\n")
                return

        mail.Send()
        self.output_console.insert(tk.END, f"E-mail odeslán na: {recipient}\n")
        time.sleep(5)  # Wait for 5 seconds



    def send_emails(self):
        # Get the directory of the current script
        script_dir = os.path.dirname(os.path.realpath(__file__))

        # Join the directory with the filename to get the full path of the file
        file_path = os.path.join(script_dir, 'mail_tool', 'recipients.xlsx')

        wb = load_workbook(file_path)
        ws = wb.active

        # Read the email list from the xlsx file
        email_list = pd.read_excel(file_path, sheet_name=ws.title)

        for _, row in email_list.iterrows():
            recipient = row['Email']
            bcc_recipient = 'martinpenkava1@gmail.com'
            attachment_name = row['Attachment']
            print(f"Připravuji mail pro {recipient}...")  # Print the recipient of the current email
            self.send_email_with_attachment(recipient, bcc_recipient, attachment_name)

        print('Emaily uspěšně odeslány.')

    def process_pdfs(self):
        # Check if 'faktury/txt' directory exists
        if os.path.exists('faktury/txt'):
            self.output_console.insert(tk.END, "Nalezl jsem informace o předchozích zpracovaných fakturách, pro jistotu je mažu, aby se Olince nestal incident :)\n")
            self.output_console.see(tk.END)
            self.output_console.update()
            shutil.rmtree('faktury/txt')

        # Create necessary directories
        os.makedirs('faktury/txt', exist_ok=True)
        os.makedirs('faktury/single_invoices', exist_ok=True)

        # Get all PDF files in the 'faktury' directory
        pdf_files = [f for f in os.listdir('faktury') if f.endswith('.pdf')]

        if not pdf_files:
            messagebox.showwarning("Olinko???", "Vrzni mi sem prosím nějakou tu fakturu, nebo rovnou dvě.")
            return

        def convert_pdf_to_txt(file_path):
            pdf_reader = PdfReader(file_path)
            texts = []
            for page in pdf_reader.pages:
                texts.append(page.extract_text())
            return texts

        def write_to_txt_and_extract_invoice_number_and_email(file_path, texts, current_invoice_number=None):
            invoice_number = current_invoice_number
            email = None
            for text in texts:
                for line in text.split('\n'):
                    if "Faktura č.:" in line:
                        match = re.search(r'Faktura č.:\s*(\d+)', line)
                        if match:
                            invoice_number = match.group(1)
                    if "E-mail:" in line:
                        match = re.search(r'(?i)E-mail:\s*([a-zA-Z0-9_.+-]+@[a-zA-Z0-9-]+\.[a-zA-Z0-9-.]+)', line)
                        if match:
                            email = match.group(1)
                    if "Podnikatel je zapsán do živnostenského rejstříku. Nejsem plátce DPH! Vystavil: Rudolf Málek" in line or "Podnikatel je zapsán do živnostenského rejstříku. Nejsem plátce DPH! Vystavil: Radim Kolečkář" in line:
                        break

            if invoice_number:
                txt_file_path = f"faktury/txt/{invoice_number}.txt"
                with open(txt_file_path, 'a', encoding='utf-8') as txt_file_obj:
                    txt_file_obj.write(text)

            return invoice_number, email

        def write_to_excel(file_path, emails, invoice_numbers):
            invoice_numbers = [str(invoice) + '.pdf' for invoice in invoice_numbers]
            df = pd.DataFrame({
                'Email': emails,
                'Attachment': invoice_numbers
            })

            if not os.path.exists('mail_tool'):
                os.makedirs('mail_tool')

            df.to_excel(file_path, index=False)
            book = load_workbook(file_path)
            sheet = book.active
            sheet.column_dimensions['A'].width = 30
            sheet.column_dimensions['B'].width = 11
            book.save(file_path)

        # Dictionary to store unique invoice-email pairs
        invoice_email_pairs = {}

        # Process each PDF file
        for pdf_file in pdf_files:
            self.output_console.insert(tk.END, f"Zpracovávám {pdf_file}...\n")
            self.output_console.see(tk.END)
            self.output_console.update()

            pdf_path = os.path.join('faktury', pdf_file)
            pdf_reader = PdfReader(pdf_path)
            
            # Variables for current invoice processing
            current_invoice_pages = []
            current_invoice_text = ""
            current_invoice_number = None
            current_email = None
            
            for page_num, page in enumerate(pdf_reader.pages):
                text = page.extract_text()
                
                # Check for invoice number in the page
                invoice_match = re.search(r'Faktura č.:\s*(\d+)', text)
                
                if invoice_match:
                    # If we have a previous invoice, save both PDF and text
                    if current_invoice_pages and current_invoice_number:
                        # Save PDF
                        output_pdf_path = f'faktury/single_invoices/{current_invoice_number}.pdf'
                        writer = PdfWriter()
                        for invoice_page in current_invoice_pages:
                            writer.add_page(invoice_page)
                        with open(output_pdf_path, 'wb') as output_file:
                            writer.write(output_file)
                        
                        # Process text for email extraction
                        invoice_number, email = write_to_txt_and_extract_invoice_number_and_email(
                            pdf_path, [current_invoice_text], current_invoice_number
                        )
                        
                        if invoice_number and email:
                            invoice_email_pairs[invoice_number] = email
                            self.output_console.insert(tk.END, f"Nalezena faktura: {invoice_number} s emailem: {email}\n")
                            self.output_console.see(tk.END)
                            self.output_console.update()
                    
                    # Start new invoice
                    current_invoice_number = invoice_match.group(1)
                    current_invoice_pages = [page]
                    current_invoice_text = text
                else:
                    # Add page to current invoice if we have one
                    if current_invoice_number:
                        current_invoice_pages.append(page)
                        current_invoice_text += "\n" + text
            
            # Process the last invoice
            if current_invoice_pages and current_invoice_number:
                # Save PDF
                output_pdf_path = f'faktury/single_invoices/{current_invoice_number}.pdf'
                writer = PdfWriter()
                for invoice_page in current_invoice_pages:
                    writer.add_page(invoice_page)
                with open(output_pdf_path, 'wb') as output_file:
                    writer.write(output_file)
                
                # Process text for email extraction
                invoice_number, email = write_to_txt_and_extract_invoice_number_and_email(
                    pdf_path, [current_invoice_text], current_invoice_number
                )
                
                if invoice_number and email:
                    invoice_email_pairs[invoice_number] = email
                    self.output_console.insert(tk.END, f"Nalezena faktura: {invoice_number} s emailem: {email}\n")
                    self.output_console.see(tk.END)
                    self.output_console.update()

            # Force garbage collection
            gc.collect()

        # Convert the dictionary to separate lists
        invoice_numbers = list(invoice_email_pairs.keys())
        emails = list(invoice_email_pairs.values())

        # Debug output
        self.output_console.insert(tk.END, f"Nalezeno unikátních faktur: {len(invoice_numbers)}\n")
        self.output_console.insert(tk.END, f"Nalezeno emailů: {len(emails)}\n")
        
        # Write to Excel
        excel_path = 'mail_tool/recipients.xlsx'
        self.output_console.insert(tk.END, "Zapisuji získaná data do recipients.xlsx pro Olinku..\n")
        self.output_console.see(tk.END)
        self.output_console.update()
        write_to_excel(excel_path, emails, invoice_numbers)
        
        def optimize_txts(directory):
            for filename in os.listdir(directory):
                if filename.endswith(".txt"):
                    with open(os.path.join(directory, filename), 'r+', encoding='utf-8') as file:
                        lines = file.readlines()
                        file.seek(0)
                        for line in lines:
                            if line.strip().endswith("Celkem"):
                                kč_index = line.rfind("Kč")
                                if kč_index != -1:
                                    print(f"Deleting from line: {line}")  # Print the line before deletion
                                    line = line[:kč_index+2] + "\n"  # +2 to include "Kč" and the space after it
                            file.write(line)
                        file.truncate()

        # Call the optimize_txts function
        optimize_txts('faktury/txt')

        self.output_console.insert(tk.END, "Zpracování dokončeno.\n")
        self.output_console.see(tk.END)
        self.output_console.update()

    def create_excel(self):
        # Path to the directory containing the script and the mail_tool folder
        dir_path = os.path.dirname(os.path.realpath(__file__))

        # Define the paths to the files
        recipients_path = os.path.join(dir_path, 'mail_tool', 'recipients.xlsx')
        build_path = os.path.join(dir_path, 'mail_tool', 'build.xlsx')

        # Check if the files exist
        if not os.path.exists(recipients_path):
            messagebox.showwarning("Olinka má průšvih!", "Soubor 'recipients.xlsx' někde Olinka zapomněla, vrzni ho do mail_tool :)")
            return  # Exit the method

        if not os.path.exists(build_path):
            messagebox.showwarning("Olinka má průšvih!", "Soubor 'build.xml' byl Olinkou zapomenut někde v řiťce. Měla bys ho rychle strčit do mail_tool :)")
            return  # Exit the method

        # Read the Excel files
        recipients_df = pd.read_excel(os.path.join(dir_path, 'mail_tool', 'recipients.xlsx'))
        build_wb = openpyxl.load_workbook(os.path.join(dir_path, 'mail_tool', 'build.xlsx'))

        # Get the unique values in the second column of the recipients file, excluding the '.pdf' part
        unique_values = recipients_df.iloc[:, 1].dropna().unique()

        # Check if there are any unique values
        if len(unique_values) == 0:
            self.output_console.insert(tk.END, "No unique values found in the second column of the recipients file.\n")
            self.output_console.see(tk.END)  # Auto-scroll to the end
            self.output_console.update()  # Ensure the output console is updated
        else:
            # Ask for the name of the new Excel file
            new_file_name = simpledialog.askstring("Zadání", "Teď by měla Olinka tento excel pojmenovat (asi podle datumu nebo jak to dělá??):")

            # Add the '.xlsx' suffix if it's not already there
            if not new_file_name.endswith('.xlsx'):
                new_file_name += '.xlsx'

            # Create a new Excel workbook
            new_wb = openpyxl.Workbook()
            new_wb.remove(new_wb.active)  # remove the default sheet created

            # Loop over the unique values
            for value in unique_values:
                sheet_name = str(value).replace('.pdf', '')
                # Copy the data from the build file to a new sheet in the new Excel file
                source = build_wb.active
                target = new_wb.create_sheet(title=sheet_name)

                for row in source:
                    for cell in row:
                        new_cell = target.cell(row=cell.row, column=cell.column, value=cell.value)
                        if cell.has_style:
                            new_cell.font = copy(cell.font)
                            new_cell.border = copy(cell.border)
                            new_cell.fill = copy(cell.fill)
                            new_cell.number_format = copy(cell.number_format)
                            new_cell.protection = copy(cell.protection)
                            new_cell.alignment = copy(cell.alignment)

                # Copy column widths
                for column in source.columns:
                    max_length = 0
                    column = [cell for cell in column]
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(cell.value)
                        except:
                            pass
                    adjusted_width = (max_length + 2)
                    target.column_dimensions[column[0].column_letter].width = adjusted_width

                # Copy row heights
                for row in source.rows:
                    for cell in row:
                        target.row_dimensions[cell.row].height = source.row_dimensions[cell.row].height

            # Save the new Excel file
            new_wb.save(os.path.join(dir_path, new_file_name))

            self.output_console.insert(tk.END, f"Otrok vytvořil {new_file_name} protože to přece nebude dělat Olinka sama ne??.\n")
            self.output_console.see(tk.END)  # Auto-scroll to the end
            self.output_console.update()  # Ensure the output console is updated
        

    def process_txts_for_plants(self):
        # Check if the 'faktury/txt' directory exists
        if not os.path.exists('faktury/txt'):
            messagebox.showwarning("Olinko???", "Nejdříve se musí pustit funkce Zpracovat Faktury, jinak ti tohle nepujde, tak to koukej pustit.")
            return  # Exit the method

        # Get all .txt files in the 'faktury/txt' directory
        txt_files = [f for f in os.listdir('faktury/txt') if f.endswith('.txt')]

        # Check if there are any .txt files
        if not txt_files:
            messagebox.showwarning("Vyhráváš zlatého Bluďišťáka", "Tohle je poměrně dobrý easter egg, protože tahle situace aby byla složka a chyběly soubory téměř nemuže nastat, tak gratuluji, že si tohle objevila, ale moc se v tom nevrtej jako!!!")
            return  # Exit the method
        
        affected_excel_files = []
        unmatched_excel_files = []

        def extract_plant_names_and_write_to_excel(file_path, invoice_number, excluded_plant_names):
            plant_names = []
            with open(file_path, 'r', encoding='utf-8') as txt_file_obj:
                lines = txt_file_obj.readlines()
                i = 0
                while i < len(lines):
                    line = lines[i].strip()
                    while line.startswith('()') and not line.endswith('Kč'):
                        i += 1
                        line += ' ' + lines[i].strip()
                    line = re.sub(r'(\d+,\d+ Kč)', r' \1', line)
                    match = re.search(r'\(\)\s(.+)', line)
                    if match:
                        plant_name = match.group(1)
                        if ("semen" in plant_name or 
                            ("hlíz" in plant_name and not ("hlíznatá" in plant_name or "hlíznatý" in plant_name)) or
                            "Dárkový poukaz" in plant_name) and "semenáč" not in plant_name and "bezsemenná" not in plant_name:
                            excluded_plant_names.append((invoice_number, plant_name))  # Add the invoice number and the excluded plant name to the list
                        else:
                            plant_names.append(plant_name)
                    i += 1

            # Write the excluded plant names to a separate text file
            with open('excluded_plant_names.txt', 'a', encoding='utf-8') as f:
                for plant_name in excluded_plant_names:
                    f.write(f'{plant_name}\n')


            new_dir = 'faktury/txt/kytky'
            if not os.path.exists(new_dir):
                os.makedirs(new_dir)

            with open(f'{new_dir}/kytky_{invoice_number}.txt', 'w', encoding='utf-8') as f:
                for plant_name in plant_names:
                    f.write(f'{plant_name}\n')

            excel_files = [f for f in os.listdir() if f.endswith('.xlsx') and f not in ['temporary.xlsx', 'template.xlsx']]

            for excel_file in excel_files:
                workbook = load_workbook(excel_file)
                if invoice_number in workbook.sheetnames:
                    worksheet = workbook[invoice_number]
                    start_row = 13
                    col = get_column_letter(3)
                    max_length = 0
                    column = col
                    for i, plant_name in enumerate(plant_names, start=start_row):
                        worksheet[f"{col}{i}"] = plant_name
                        if len(plant_name) > max_length:
                            max_length = len(plant_name)
                    worksheet.column_dimensions[column].width = max_length
                    workbook.save(excel_file)
                    if excel_file not in affected_excel_files:
                        affected_excel_files.append(excel_file)
                else:
                    if excel_file not in unmatched_excel_files:
                        unmatched_excel_files.append(excel_file)


        # Process each .txt file
        excluded_plant_names = []  # Define the list outside of the function
        for txt_file in txt_files:
            # Skip 'kytky_plant_names.txt' and 'plant_names.txt'
            if txt_file in ['kytky_plant_names.txt', 'plant_names.txt']:
                continue

            self.output_console.insert(tk.END, f"Zpracovávám {txt_file}...\n")
            self.output_console.see(tk.END)  # Auto-scroll to the end
            self.output_console.update()  # Ensure the output console is updated

            invoice_number = os.path.splitext(txt_file)[0].replace('kytky_', '')  # Get the invoice number from the file name
            txt_path = os.path.join('faktury/txt', txt_file)
            extract_plant_names_and_write_to_excel(txt_path, invoice_number, excluded_plant_names)

        self.output_console.insert(tk.END, "Zpracování dokončeno.\n")
        self.output_console.insert(tk.END, f"Excely které jsem naplnil rostlinami: {', '.join(affected_excel_files)}\n")
        self.output_console.insert(tk.END, f"Excely obsahující jiné faktury a tudíž nebyly naplněny rostlinami: {', '.join(unmatched_excel_files)}\n")
        self.output_console.see(tk.END)  # Auto-scroll to the end
        self.output_console.update()  # Ensure the output console is updated


        # Write the excluded plant names to a separate text file and Excel file
        with open('excluded_plant_names.txt', 'a', encoding='utf-8') as f:
            wb = Workbook()
            ws = wb.active
            ws.append(["Číslo faktury", "Název rostliny"])  # Add the header to the Excel file
            for invoice_number, plant_name in excluded_plant_names:
                f.write(f'{invoice_number}: {plant_name}\n')
                ws.append([invoice_number, plant_name])  # Append the invoice number and the plant name to the Excel file

            ws.auto_filter.ref = ws.dimensions  # Add autofilter to the worksheet

            # Adjust column width
            for column in ws.columns:
                max_length = 0
                column = [cell for cell in column]
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(cell.value)
                    except:
                        pass
                adjusted_width = (max_length + 2)
                ws.column_dimensions[column[0].column_letter].width = adjusted_width

            wb.save("faktury/excluded_plant_names.xlsx")

        os.remove('excluded_plant_names.txt')  # Delete the text file

        # Ask the user if they want to open the created Excel file
        if messagebox.askyesno("Otevřít soubor", "Moje milá Olinko, máš to hotové a protože jsi to ty a já ti chtěl udělat radost, vytvořil jsem ti soubor, kde je zaznamenáno veškeré semeno a hlíza a taky DÁRKOVÝ POUKAZ které jsem dal pryč, chtěla bys tam nahlédnout?"):
            # Get the directory of the current script
            script_dir = os.path.dirname(os.path.realpath(__file__))

            # Construct the path to the file
            file_path = os.path.join(script_dir, 'faktury', 'excluded_plant_names.xlsx')

            # Open the file
            os.startfile(file_path)

    def open_trim_names_window(self):
        new_window = Toplevel(self.master)
        new_window.title("Ořezat názvy")
        new_window.geometry("1230x900")  # Set the size of the window

        listbox_widget = tk.Listbox(new_window, height=35, width=200)
        listbox_widget.grid(row=0, column=0, columnspan=3, padx=10, pady=10)

        # Add instructions label
        instructions = "Použití:\n" \
                    "1. Nejdřív načti Excel - kliknutím na vybrat Excel.\n" \
                    "2. Potom můžeš zobrazit, jaký změny tool dokáže udělat kliknutím na 'Zobrazit, co se změní'\n" \
                    "3. Kliknutím na Zpracovat Excely provedeš změny, které byly vidět v předchozím kroku \n" \
                    "4. Funkce zobrazit, co ještě zbývá změnit zobrazuje hodnoty, které nebyly změněny toolem - doporučeno spustit až po prvním zpracování excelu.\n" \
                    "5. Pokud načteš slovník a něco se ti v něm nelíbí je možné hodnotu smazat přímo přes tool zvolením krátkého názvu a stiskuntím klávesy ''Delete''\n" \
        
        instructions_label = tk.Label(new_window, text=instructions)
        instructions_label.grid(row=1, column=0, columnspan=3, padx=10, pady=10)

        entry_widget = tk.Entry(new_window, width=75)
        entry_widget.grid(row=2, column=0, columnspan=3, padx=10, pady=10)

        listbox_widget = tk.Listbox(new_window, height=35, width=200)
        listbox_widget.grid(row=0, column=0, columnspan=3, padx=10, pady=10)
        listbox_widget.bind('<<ListboxSelect>>', lambda event: self.on_listbox_select(event, entry_widget))
        listbox_widget.bind('<Delete>', lambda event: self.on_delete_click(listbox_widget))

        show_changes_button = Button(new_window, text="Zobrazit, co se změní", command=lambda: self.on_show_changes_click(listbox_widget))
        show_changes_button.grid(row=2, column=0, padx=10, pady=10)

        self.show_unmatched_button = Button(new_window, text="Zobrazit, co ještě zbývá změnit", command=lambda: self.on_show_unmatched_click(listbox_widget), state='disabled')
        self.show_unmatched_button.grid(row=3, column=0, padx=10, pady=10)

        process_button = Button(new_window, text="Zpracovat Excel", command=lambda: self.on_process_click(listbox_widget))
        process_button.grid(row=4, column=0, padx=10, pady=10)

        load_button = Button(new_window, text="Zobrazit slovník", command=lambda: self.on_load_click(listbox_widget))
        load_button.grid(row=2, column=2, padx=10, pady=10)

        add_button = Button(new_window, text="Přidat do slovníku", command=lambda: self.on_add_click(entry_widget))
        add_button.grid(row=3, column=2, padx=10, pady=10)

        select_file_button = Button(new_window, text="Vybrat Excel", command=self.select_file)
        select_file_button.grid(row=6, column=1, padx=10, pady=10)


        load_from_template_button = Button(new_window, text="Načíst slovník ze šablony", command=lambda: self.load_from_template(listbox_widget))
        load_from_template_button.grid(row=4, column=2, padx=10, pady=10)




        # Automatically process the Excel file and display the unmatched values
        #self.on_process_click(listbox_widget)


    def load_dictionary(self):
        with open('dictionary.txt', 'r', encoding='utf-8') as f:
            return f.read().splitlines()
    
    def select_file(self):
        self.excel_file = filedialog.askopenfilename(parent=self.master, filetypes=[("Excel files", "*.xlsx")])
 
    def process_excel(self, dictionary_values, show_changes=False):
        dictionary_values.sort(key=len, reverse=True)
        if not hasattr(self, 'excel_file'):
            messagebox.showerror("Chyba! Ajéje", "Olinka musí vybrat nejdřív excel přes tlačítko Vybrat Excel, to je logický ne??")
            return

        book = load_workbook(self.excel_file)
        values = []

        for sheet_name in book.sheetnames:
            sheet = book[sheet_name]
            for row in range(13, sheet.max_row + 1):
                cell = sheet.cell(row=row, column=3)
                if cell.value and cell.value not in dictionary_values:  # check if the cell value is not in the dictionary
                    original_value = cell.value  # store the original value
                    for value in dictionary_values:
                        if value in original_value:
                            new_value = value  # store the new value
                            if not show_changes:  # only print to console if show_changes is False
                                self.output_console.insert(tk.END, f'Přepsáno z: {original_value} >> Nový: {new_value}\n')
                                self.output_console.see(tk.END)
                            if show_changes:
                                values.append((original_value, new_value))  # use the stored original value and the new value
                            else:
                                cell.value = new_value  # replace the cell value with the new value
                            break
                    if not show_changes and cell.value == original_value:
                        values.append((original_value, 'Unmatched'))

        if not show_changes:
            book.save(self.excel_file)  # only save the changes to the Excel file when show_changes is False

        return values





    def add_to_dictionary(self, value):
        with open('dictionary.txt', 'a+', encoding='utf-8') as f:
            # Move the pointer to the start of the file
            f.seek(0)
            # Check if the file is not empty
            if f.read(1):
                # If the file is not empty, prepend the value with a newline character
                f.write(f'\n{value}')
            else:
                # If the file is empty, just write the value
                f.write(value)



    def on_process_click(self, listbox_widget):
        dictionary_values = self.load_dictionary()
        unmatched_values = self.process_excel(dictionary_values)
        listbox_widget.delete(0, tk.END)
        for original_value, new_value in unmatched_values:
            if new_value == 'Unmatched':
                listbox_widget.insert(tk.END, original_value)

        self.show_unmatched_button.config(state='normal')  # enable the button



    def on_load_click(self, listbox_widget):
        dictionary_values = self.load_dictionary()
        listbox_widget.delete(0, tk.END)
        for value in dictionary_values:
            listbox_widget.insert(tk.END, value)


    def on_show_changes_click(self, listbox_widget):
        dictionary_values = self.load_dictionary()
        values = self.process_excel(dictionary_values, show_changes=True)
        listbox_widget.delete(0, tk.END)
        if not values:  # if no changes were made
            listbox_widget.insert(tk.END, "Nenašel jsem přepisy | Nic přepsáno pro Olinku nebude!")
        else:
            for original_value, new_value in values:
                listbox_widget.insert(tk.END, f'Originál: {original_value} >> Nový: {new_value}')

    def on_add_click(self, entry_widget):
        value = entry_widget.get()
        if value:
            self.add_to_dictionary(value)
            self.output_console.insert(tk.END, f"Povedlo se!: Název '{value}' byl přidán do slovníku\n")

    def show_unmatched(self, dictionary_values):
        if not hasattr(self, 'excel_file'):
            messagebox.showerror("Error", "No file selected")
            return

        book = load_workbook(self.excel_file, read_only=True)
        unmatched_values = []

        for sheet_name in book.sheetnames:
            sheet = book[sheet_name]
            for row in range(13, sheet.max_row + 1):
                cell = sheet.cell(row=row, column=3)
                if cell.value and cell.value not in dictionary_values:  # check if the cell value is not in the dictionary
                    unmatched_values.append(cell.value)

        return unmatched_values

    def on_show_unmatched_click(self, listbox_widget):
        dictionary_values = self.load_dictionary()
        unmatched_values = self.show_unmatched(dictionary_values)
        listbox_widget.delete(0, tk.END)
        if unmatched_values:
            for value in unmatched_values:
                listbox_widget.insert(tk.END, value)
        else:
            listbox_widget.insert(tk.END, "Všechno už bylo zpracováno, tento Excel už obsahuje pouze latinské názvy :)")


    def on_listbox_select(self, event, entry_widget):
        # Get selected line index
        index = event.widget.curselection()
        # Check if the index is not empty
        if index:
            # Get the line's text
            selected_text = event.widget.get(index)
            # Clear the entry widget
            entry_widget.delete(0, tk.END)
            # Insert the selected text into the entry widget
            entry_widget.insert(0, selected_text)


    def update_listbox_with_dictionary(self, listbox_widget):
        dictionary_values = self.load_dictionary()
        listbox_widget.delete(0, tk.END)
        for value in dictionary_values:
            listbox_widget.insert(tk.END, value)

    def on_delete_click(self, listbox_widget):
        # Get selected line index
        index = listbox_widget.curselection()
        # Check if the index is not empty
        if index:
            # Get the line's text
            selected_text = listbox_widget.get(index)
            # Delete the selected line from the listbox
            listbox_widget.delete(index)
            # Delete the selected line from the dictionary.txt file
            self.delete_from_dictionary(selected_text)

    def delete_from_dictionary(self, value):
        # Load all dictionary values
        dictionary_values = self.load_dictionary()
        # Remove the selected value
        dictionary_values.remove(value)
        # Write the updated dictionary values back to the file
        with open('dictionary.txt', 'w', encoding='utf-8') as f:
            f.write('\n'.join(dictionary_values))


    def load_from_template(self, listbox_widget):
        # Load the workbook
        book = load_workbook('template.xlsx')
        # Select the first sheet
        sheet = book.active
        # Get the values from column A starting from row 2
        values = [cell.value for cell in sheet['A'][1:] if cell.value is not None]
        # Write the values to the dictionary.txt file
        with open('dictionary.txt', 'w', encoding='utf-8') as f:
            f.write('\n'.join(values))
        # Update the listbox with the new dictionary values
        self.update_listbox_with_dictionary(listbox_widget)

    def remove_duplicates(self):
        duplicates_found = False

        for filename in os.listdir('.'):
            if not filename.endswith('.xlsx') or filename == 'template.xlsx' or filename == 'temporary.xlsx':
                continue

            wb = openpyxl.load_workbook(filename)

            for sheet in wb:
                ws = wb[sheet.title]
                unique_values = set()

                # Create a new sheet to store the non-duplicate rows
                new_ws = wb.create_sheet(title=f"{ws.title}_no_duplicates")
                
                new_row = 1  # Counter for the new sheet's row
                for row in range(1, ws.max_row + 1):  # Start from 1 to include headers
                    value = ws.cell(row=row, column=3).value
                    
                    if row >= 13 and value and value in unique_values:
                        self.output_console.insert(tk.END, f"Nalezena duplicita: {value} v souboru: {filename}, sheet: {sheet.title}, řádek: {row}\n")
                        self.output_console.see(tk.END)
                        self.output_console.update()
                        duplicates_found = True
                        continue  # Skip this row, don't copy it to the new sheet

                    # If value is valid and not a duplicate, add it to the set and copy the row to the new sheet
                    if row < 13 or value:  # Always copy headers
                        if value:
                            unique_values.add(value)
                        for col in range(1, ws.max_column + 1):  # Copy all columns
                            new_cell = new_ws.cell(row=new_row, column=col)
                            old_cell = ws.cell(row=row, column=col)
                            # Copy style attributes
                            new_cell.font = copy(old_cell.font)
                            new_cell.border = copy(old_cell.border)
                            new_cell.fill = copy(old_cell.fill)
                            new_cell.number_format = copy(old_cell.number_format)
                            new_cell.protection = copy(old_cell.protection)
                            new_cell.alignment = copy(old_cell.alignment)
                            new_cell.value = old_cell.value
                        new_row += 1  # Move to the next row in the new sheet

                # Delete the old sheet and rename the new one
                wb.remove(ws)
                new_ws.title = sheet.title

            wb.save(filename)

            if not duplicates_found:
                self.output_console.insert(tk.END, "Olinka tu nemá žádné duplicity, všechno je v oukeji.\n")
                self.output_console.see(tk.END)
                self.output_console.update()

        # Your additional script starts here
        directory = os.getcwd()

        for filename in os.listdir(directory):
            if not filename.endswith('.xlsx') or filename in ['temporary.xlsx', 'template.xlsx']:
                continue

            filepath = os.path.join(directory, filename)
            wb = openpyxl.load_workbook(filepath)

            for sheet in wb.sheetnames:
                if not wb[sheet]['C13'].value:
                    self.output_console.insert(tk.END, f"Mažu sheet: {sheet} z: {filename} protože píčus zákazník nakoupil jen hlízu nebo semeno a nebo ještě si ten blbec koupil jen dárkový poukaz, ale nic pro sebe a mně to kvuli němu nefungovalo a sheet by byl tedy prázdný a Olinka by posílala maily s prázdnými pasy.\n")
                    self.output_console.see(tk.END)
                    self.output_console.update()
                    del wb[sheet]

            wb.save(filepath)
        # Your additional script ends here

    def open_tropik(self):
        webbrowser.open('https://www.tropik.cz')

    def open_eshop_tropik(self):
        webbrowser.open('https://eshop.tropik.cz')
    

    def quit(self):
        self.master.destroy()

if __name__ == "__main__":
    root = tk.Tk()
    app = PlantCodeFinder(root)
    root.resizable(False, False)  # Disable maximizing
    root.mainloop()

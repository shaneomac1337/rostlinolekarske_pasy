import os
import pandas as pd
import openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font, Border, Side, Alignment

def compress_excel_file(input_file, output_file):
    # Load spreadsheet
    xl = pd.ExcelFile(input_file)

    # Load a sheet into a DataFrame by its name
    df_dict = {sheet_name: xl.parse(sheet_name) for sheet_name in xl.sheet_names}

    # Create a new workbook
    wb = openpyxl.Workbook()

    for sheet_name, df in df_dict.items():
        # Create a new sheet
        ws = wb.create_sheet(title=sheet_name)

        # Write DataFrame to worksheet
        for r in dataframe_to_rows(df, index=False, header=True):
            ws.append(r)

        # Delete cells A1-E1
        for col in ['A', 'B', 'C', 'D', 'E']:
            ws[col + '1'].value = None

        # Merge cells D7-E7, D8-E8, and C6-C10
        ws.merge_cells('D7:E7')
        ws.merge_cells('D8:E8')
        ws.merge_cells('C6:C10')

        # Set font to bold and size to 14 for cells D7 and D8
        bold_font = Font(bold=True, size=14)
        ws['D7'].font = bold_font
        ws['D8'].font = bold_font

        # Set column widths
        ws.column_dimensions['B'].width = 4.71
        ws.column_dimensions['C'].width = 48.71
        ws.column_dimensions['D'].width = 21.14
        ws.column_dimensions['E'].width = 23.85
        ws.column_dimensions['F'].width = 4.28

        # Define border styles
        left_border = Border(left=Side(style='thin'))
        right_border = Border(right=Side(style='thin'))
        top_border = Border(top=Side(style='thin'))
        bottom_border = Border(bottom=Side(style='thin'))
        inside_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

        # Determine the last row with data in column C
        last_row = max(31, max((cell.row for cell in ws['C'] if cell.value is not None)) + 1)

        # Apply borders to the range B5:F31 (or B5:F32 if there's data in C31)
        for row in ws.iter_rows(min_row=5, max_row=last_row, min_col=2, max_col=6):
            for cell in row:
                if cell.row == 5:
                    cell.border += top_border
                if cell.row == last_row:
                    cell.border += bottom_border
                if cell.column == 2:  # Column B
                    cell.border += left_border
                if cell.column == 6:  # Column F
                    cell.border += right_border

        # Apply inside borders to the range C12:E30 (or further if there's data in C31 or beyond)
        for row in ws.iter_rows(min_row=12, max_row=last_row-1, min_col=3, max_col=5):
            for cell in row:
                cell.border = inside_border

        # Make C12, D12, E12 align on center and bold
        center_aligned_bold_font = Font(bold=True)
        center_alignment = Alignment(horizontal='center')
        for col in ['C', 'D', 'E']:
            cell = ws[col + '12']
            cell.font = center_aligned_bold_font
            cell.alignment = center_alignment

        # Change the font of C13 and further down, D13 and further down, and E13 and further down to Arial 11
        arial_11_font = Font(name='Arial', size=10)
        for col in ['C', 'D', 'E']:
            for row in range(13, last_row):
                cell = ws[col + str(row)]
                cell.font = arial_11_font

    # Delete the default sheet created and save the workbook
    del wb['Sheet']
    wb.save(output_file)

# Get all Excel files in the same directory as the script, excluding 'template.xlsx' and 'temporary.xlsx'
excel_files = [f for f in os.listdir() if f.endswith('.xlsx') and f not in ['template.xlsx', 'temporary.xlsx']]

# Use the function for each file
for file in excel_files:
    compress_excel_file(file, 'compressed_' + file)